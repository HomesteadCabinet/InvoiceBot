from __future__ import annotations

from datetime import datetime, timedelta, date
from decimal import Decimal, InvalidOperation
from email.utils import parsedate_to_datetime
from io import BytesIO
import logging
import os
import re
import threading
import time

from django.conf import settings
from django.db import transaction
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook

from . import parsers as parser_module
from .models import (
    Contact,
    EmailMessageCache,
    Invoice,
    InvoiceAutomationSettings,
    InventoryItem,
    Job,
    LineItem,
    ItemType,
    ProcessedEmail,
    Vendor,
    VendorEmail,
    exclude_ignored_vendor_relations,
)
from .item_types import resolve_item_type
from .parsers import list_invoice_parsers, normalize_parser_output
from .utils import get_gmail_service

logger = logging.getLogger(__name__)

_worker_thread = None
_worker_lock = threading.Lock()
_processing_lock = threading.Lock()

GMAIL_INVOICE_QUERY = 'has:attachment invoice'


def media_url_for_stored_filename(stored_filename):
    """Relative URL for saved PDFs (works with Vite /api and /media proxies)."""
    media_prefix = settings.MEDIA_URL.strip('/')
    return f'/{media_prefix}/{stored_filename}'


def attachment_info_from_cache(cache):
    if not cache or not cache.attachment_filename:
        return None
    return {
        'filename': cache.attachment_filename,
        'original_filename': cache.attachment_original_filename or cache.attachment_filename,
        'mimeType': cache.attachment_mime_type or 'application/pdf',
        'url': media_url_for_stored_filename(cache.attachment_filename),
    }


def _update_email_cache_attachment(message_id, attachment_info):
    if not attachment_info:
        return None
    cache, _created = EmailMessageCache.objects.update_or_create(
        email_id=message_id,
        defaults={
            'attachment_filename': attachment_info.get('filename') or '',
            'attachment_original_filename': attachment_info.get('original_filename') or '',
            'attachment_mime_type': attachment_info.get('mimeType') or 'application/pdf',
            'last_seen_at': timezone.now(),
        },
    )
    return cache


def _safe_filename_part(value, fallback='invoice'):
    text = str(value or '').strip()
    if not text:
        text = fallback
    text = re.sub(r'[^a-zA-Z0-9._-]+', '_', text)
    text = re.sub(r'_+', '_', text).strip('._-')
    return text or fallback


def _unique_media_filename(media_dir, filename):
    base, extension = os.path.splitext(filename)
    candidate = filename
    index = 2
    while os.path.exists(os.path.join(media_dir, candidate)):
        candidate = f'{base}_{index}{extension}'
        index += 1
    return candidate


def _invoice_job_or_po(invoice):
    line_items = invoice.line_items.select_related('job').all()
    job_values = sorted({
        ' '.join(part for part in (
            line_item.job.job_id,
            line_item.job.name,
        ) if part)
        for line_item in line_items
        if line_item.job_id
    })
    if len(job_values) == 1:
        return job_values[0]
    return invoice.customer_po or ''


def _attachment_filename_for_invoices(message_id, vendor, invoices, original_filename):
    extension = os.path.splitext(original_filename or '')[1] or '.pdf'
    vendor_name = (
        vendor.name
        if vendor
        else next((invoice.vendor.name for invoice in invoices if invoice.vendor_id), '')
    )
    po_or_job_values = [
        _invoice_job_or_po(invoice)
        for invoice in invoices
        if _invoice_job_or_po(invoice)
    ]
    unique_po_or_job_values = list(dict.fromkeys(po_or_job_values))
    if len(unique_po_or_job_values) == 1:
        descriptor = unique_po_or_job_values[0]
    elif len(unique_po_or_job_values) > 1:
        descriptor = 'multiple_jobs'
    else:
        descriptor = 'invoice'

    return (
        f'{_safe_filename_part(message_id)}_'
        f'{_safe_filename_part(vendor_name, "vendor")}_'
        f'{_safe_filename_part(descriptor)}'
        f'{extension.lower()}'
    )


def _rename_attachment_for_invoices(file_path, message_id, vendor, invoices, original_filename):
    if not invoices:
        return os.path.basename(file_path)
    media_dir = os.path.dirname(file_path)
    target_filename = _attachment_filename_for_invoices(message_id, vendor, invoices, original_filename)
    target_filename = _unique_media_filename(media_dir, target_filename)
    target_path = os.path.join(media_dir, target_filename)
    if os.path.abspath(file_path) != os.path.abspath(target_path):
        os.replace(file_path, target_path)
    return target_filename


def attachment_info_for_message(message_id):
    """Find a previously saved PDF for this Gmail message id."""
    cached = attachment_info_from_cache(
        EmailMessageCache.objects.filter(email_id=message_id).first()
    )
    if cached:
        return cached

    media_dir = settings.MEDIA_ROOT
    if not os.path.isdir(media_dir):
        return None

    prefix = f'{message_id}_'
    for name in sorted(os.listdir(media_dir)):
        if not name.startswith(prefix):
            continue
        if not name.lower().endswith('.pdf'):
            continue
        attachment_info = {
            'filename': name,
            'original_filename': name[len(prefix):],
            'mimeType': 'application/pdf',
            'url': media_url_for_stored_filename(name),
        }
        _update_email_cache_attachment(message_id, attachment_info)
        return attachment_info
    return None


def _extract_sender_email(from_header):
    if not from_header:
        return None
    email_match = re.search(r'<(.+?)>|([^<\s]+@[^>\s]+)', from_header)
    if not email_match:
        return None
    return email_match.group(1) or email_match.group(2)


def _vendor_name_from_sender(from_header, sender_email):
    if not sender_email:
        return None
    special_domains = (
        'notification.intuit.com',
        'billtrust.com',
        'live.com',
        'outlook.com',
        'gmail.com',
        'yahoo.com',
        'hotmail.com',
        'msn.com',
    )
    if any(sender_email.endswith(f'@{domain}') for domain in special_domains):
        name_match = re.search(r'^(.+?)\s*<', from_header or '')
        if name_match:
            return name_match.group(1).strip()
    domain_match = re.search(r'@(.+?)\.[^.]+$', sender_email)
    return domain_match.group(1).title() if domain_match else 'Unknown'


def _sync_vendor_for_sender(from_header, sender_email):
    vendor_name = _vendor_name_from_sender(from_header, sender_email)
    if not vendor_name:
        return None
    vendor, _created = Vendor.objects.update_or_create(
        name=vendor_name,
        defaults={'invoice_type': 'pdf'},
    )
    VendorEmail.objects.update_or_create(
        email=sender_email,
        defaults={'vendor': vendor, 'is_primary': True},
    )
    return vendor


def vendor_is_ignored(vendor):
    return bool(vendor and vendor.ignore)


def _parse_date(value):
    if not value:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    for fmt in ('%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d'):
        try:
            return datetime.strptime(str(value), fmt).date()
        except ValueError:
            continue
    return None


def _parse_datetime(value):
    if not value:
        return None
    if isinstance(value, datetime):
        return value
    try:
        return parsedate_to_datetime(str(value))
    except Exception:
        return None


def _decimal(value):
    if value is None or value == '':
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _decimal_for_field(value, *, max_digits, decimal_places):
    """
    Parse ``value`` for a DecimalField(max_digits, decimal_places).

    Returns None when input is empty, unparseable, or outside field limits.
    """
    parsed = _decimal(value)
    if parsed is None:
        return None
    quantum = Decimal(10) ** -decimal_places
    max_abs = Decimal(10) ** (max_digits - decimal_places) - quantum
    if parsed.copy_abs() > max_abs:
        logger.warning(
            'Ignoring out-of-range decimal %r (max abs %s for %s,%s)',
            value,
            max_abs,
            max_digits,
            decimal_places,
        )
        return None
    try:
        return parsed.quantize(quantum)
    except InvalidOperation:
        logger.warning('Ignoring non-quantizable decimal %r', value)
        return None


def _decimal_12_4(value):
    return _decimal_for_field(value, max_digits=12, decimal_places=4)


def _decimal_12_2(value):
    return _decimal_for_field(value, max_digits=12, decimal_places=2)


def _line_item_qty(value):
    parsed = _decimal_12_4(value)
    if parsed is not None:
        return parsed
    if value not in (None, '') and _decimal(value) is not None:
        logger.warning('Line item qty out of range: %r; using 0', value)
    return Decimal('0')


def _selected_parser_for_vendor(vendor):
    if not vendor:
        return None

    if vendor.parser:
        return getattr(parser_module, vendor.parser, None)
    vendor_name = (vendor.name or '').lower()
    if 'sherwin' in vendor_name:
        return getattr(parser_module, 'parse_sherwin_invoice', None)
    return getattr(parser_module, 'parse_generic_invoice', None)


def _ensure_invoice_automation_settings():
    return InvoiceAutomationSettings.load()


def resolve_vendor_for_persist(vendor=None, parsed_output=None, email_payload=None):
    """Resolve vendor from explicit instance, parsed PDF data, or sender email."""
    if vendor is not None:
        return vendor
    vendor_name = (parsed_output or {}).get('vendor_name')
    if vendor_name:
        vendor, _ = Vendor.objects.get_or_create(
            name=str(vendor_name).strip(),
            defaults={'invoice_type': 'pdf'},
        )
        return vendor
    email_payload = email_payload or {}
    from_header = email_payload.get('from') or ''
    sender_email = _extract_sender_email(from_header)
    if sender_email:
        return _sync_vendor_for_sender(from_header, sender_email)
    return None


def resolve_contact(vendor, email_payload):
    """Get or create a contact from invoice email headers."""
    if not vendor:
        return None
    email_payload = email_payload or {}
    from_header = str(email_payload.get('from') or '').strip()
    sender_email = (_extract_sender_email(from_header) or '').strip()
    name = ''
    if from_header:
        name_match = re.search(r'^(.+?)\s*<', from_header)
        if name_match:
            name = name_match.group(1).strip().strip('"')
    if not name and sender_email:
        name = sender_email.split('@')[0].replace('.', ' ').title()
    if not sender_email and not name:
        return vendor.contacts.filter(is_primary=True).first()

    if sender_email:
        contact = Contact.objects.filter(vendor=vendor, email=sender_email).first()
        if not contact:
            contact = Contact.objects.create(
                vendor=vendor,
                email=sender_email,
                name=name or sender_email,
                is_primary=False,
            )
    else:
        contact = Contact.objects.filter(vendor=vendor, name=name, email='').first()
        if not contact:
            contact = Contact.objects.create(
                vendor=vendor,
                name=name,
                email='',
                is_primary=False,
            )
    if name and contact.name != name:
        contact.name = name
        contact.save(update_fields=['name'])
    return contact


def resolve_job(vendor, job_id='', job_name=''):
    """Get or create a Job from parsed line-item job id and name."""
    job_id = str(job_id or '').strip()
    job_name = str(job_name or '').strip()
    if not job_id and not job_name:
        return None
    if job_id:
        job, _created = Job.objects.update_or_create(
            vendor=vendor,
            job_id=job_id,
            defaults={'name': job_name},
        )
        return job
    job, _ = Job.objects.get_or_create(
        vendor=vendor,
        job_id='',
        name=job_name,
    )
    return job


def _normalize_inventory_key_value(value):
    if value is None:
        return ''
    return str(value).strip().casefold()


def _inventory_item_key(line_item_payload):
    name = str(line_item_payload.get('name') or '').strip()
    if name:
        return _normalize_inventory_key_value(name)
    item_id = str(line_item_payload.get('id') or line_item_payload.get('item_id') or '').strip()
    if item_id:
        return _normalize_inventory_key_value(item_id)
    return ''


def _inventory_item_label(line_item_payload):
    name = str(line_item_payload.get('name') or '').strip()
    if name:
        return name
    item_id = str(line_item_payload.get('id') or line_item_payload.get('item_id') or '').strip()
    return item_id


def _revert_inventory_from_existing_invoice(invoice):
    for line_item in invoice.line_items.select_related('inventory_item').all():
        inventory_item = line_item.inventory_item
        if not inventory_item:
            continue
        qty = _decimal_12_4(line_item.qty) or Decimal('0')
        inventory_item.current_qty = (inventory_item.current_qty or Decimal('0')) - qty
        inventory_item.last_invoiced_at = invoice.processed_at or inventory_item.last_invoiced_at
        inventory_item.save(update_fields=['current_qty', 'last_invoiced_at', 'updated_at'])


def sync_invoice_receipt_status(invoice):
    line_items = invoice.line_items.all()
    if not line_items.exists():
        return invoice

    any_received = line_items.filter(received=True).exists()
    all_received = not line_items.filter(received=False).exists()

    if all_received:
        next_status = 'received'
    elif any_received:
        next_status = 'partially_received'
    else:
        next_status = 'processed'

    if invoice.status != next_status:
        invoice.status = next_status
        invoice.save(update_fields=['status', 'updated_at'])
    return invoice


def _update_inventory_from_line_item(invoice, line_item, line_item_payload):
    item_key = _inventory_item_key(line_item_payload)
    if not item_key:
        return None
    current_qty = _decimal_12_4(line_item_payload.get('qty'))
    if current_qty is None:
        if line_item_payload.get('qty') not in (None, '') and _decimal(line_item_payload.get('qty')) is not None:
            logger.warning(
                'Skipping inventory update for %r: qty out of range',
                _inventory_item_key(line_item_payload),
            )
            return None
        current_qty = Decimal('0')
    inventory_item, _created = InventoryItem.objects.get_or_create(
        vendor=invoice.vendor,
        item_key=item_key,
        defaults={
            'item_type': line_item.item_type,
            'item_id': str(line_item_payload.get('id') or ''),
            'name': _inventory_item_label(line_item_payload),
            'description': str(line_item_payload.get('description') or ''),
            'unit': str(line_item_payload.get('unit') or ''),
            'current_qty': current_qty,
            'last_unit_price': _decimal_12_4(line_item_payload.get('unit_price')),
            'last_total_price': _decimal_12_4(line_item_payload.get('total_price')),
            'last_invoiced_at': invoice.processed_at or timezone.now(),
            'metadata': {'last_invoice_id': invoice.id},
        },
    )
    if not _created:
        inventory_item.item_type = line_item.item_type or inventory_item.item_type
        inventory_item.item_id = str(line_item_payload.get('id') or inventory_item.item_id)
        inventory_item.name = _inventory_item_label(line_item_payload) or inventory_item.name
        inventory_item.description = str(line_item_payload.get('description') or inventory_item.description)
        inventory_item.unit = str(line_item_payload.get('unit') or inventory_item.unit)
        inventory_item.current_qty = (inventory_item.current_qty or Decimal('0')) + current_qty
        inventory_item.last_unit_price = _decimal_12_4(line_item_payload.get('unit_price'))
        inventory_item.last_total_price = _decimal_12_4(line_item_payload.get('total_price'))
        inventory_item.last_invoiced_at = invoice.processed_at or timezone.now()
        inventory_item.metadata = {**(inventory_item.metadata or {}), 'last_invoice_id': invoice.id}
        inventory_item.save(update_fields=[
            'item_type', 'item_id', 'name', 'description', 'unit', 'current_qty',
            'last_unit_price', 'last_total_price', 'last_invoiced_at', 'metadata', 'updated_at'
        ])
    line_item.inventory_item = inventory_item
    line_item.save(update_fields=['inventory_item', 'updated_at'])
    return inventory_item


def _normalize_line_item_key_value(value):
    if value is None:
        return ''
    return str(value).strip().lower()


def _first_payload_value(payload, *keys):
    for key in keys:
        value = payload.get(key)
        if value not in (None, ''):
            return value
    return ''


def _invoice_customer_po(invoice_payload):
    return str(_first_payload_value(
        invoice_payload,
        'cust_po',
        'customer_po',
        'customerPO',
        'po',
        'po_number',
        'poNumber',
        'Customer PO',
        'PO Number',
        'P.O. Number',
    ) or '').strip()


def _line_item_job_id(line_item_payload):
    return str(_first_payload_value(
        line_item_payload,
        'job_id',
        'job_number',
        'jobNumber',
        'job_no',
        'jobNo',
        'Job ID',
        'Job Number',
    ) or '').strip()


def _line_item_job_name(line_item_payload):
    return str(_first_payload_value(
        line_item_payload,
        'job',
        'job_name',
        'jobName',
        'Job',
        'Job Name',
    ) or '').strip()


def _decimal_key_value(value):
    decimal_value = _decimal_12_4(value)
    if decimal_value is None:
        return ''
    return format(decimal_value.normalize(), 'f')


def _line_item_state_key(line_item_payload):
    return (
        _normalize_line_item_key_value(line_item_payload.get('id') or line_item_payload.get('item_id')),
        _normalize_line_item_key_value(line_item_payload.get('name')),
        _normalize_line_item_key_value(line_item_payload.get('description')),
        _normalize_line_item_key_value(_line_item_job_id(line_item_payload)),
        _normalize_line_item_key_value(_line_item_job_name(line_item_payload)),
        _decimal_key_value(line_item_payload.get('qty')),
        _normalize_line_item_key_value(line_item_payload.get('unit')),
        _decimal_key_value(line_item_payload.get('unit_price')),
        _decimal_key_value(line_item_payload.get('total_price')),
        _decimal_key_value(line_item_payload.get('width')),
        _decimal_key_value(line_item_payload.get('length')),
        _decimal_key_value(line_item_payload.get('height')),
    )


def _line_item_state_map(invoice):
    state_map = {}
    for line_item in invoice.line_items.all():
        state_map.setdefault(_line_item_state_key({
            'id': line_item.item_id,
            'name': line_item.name,
            'description': line_item.description,
            'job_id': line_item.job.job_id if line_item.job_id else '',
            'job': line_item.job.name if line_item.job_id else '',
            'qty': line_item.qty,
            'unit': line_item.unit,
            'unit_price': line_item.unit_price,
            'total_price': line_item.total_price,
            'width': line_item.width,
            'length': line_item.length,
            'height': line_item.height,
        }), []).append({
            'received': line_item.received,
            'notes': line_item.notes,
        })
    return state_map


def _create_line_items_for_invoice(invoice, vendor, invoice_payload, existing_state=None):
    """Create LineItem, Job, ItemType, and InventoryItem rows from parser output."""
    existing_state = existing_state or {}
    for line_item_payload in invoice_payload.get('line_items', []) or []:
        item_type_name = str(
            line_item_payload.get('item_type') or line_item_payload.get('type') or ''
        ).strip()
        item_type = resolve_item_type(item_type_name) if item_type_name else None
        state_key = _line_item_state_key(line_item_payload)
        preserved_state = existing_state.get(state_key, [])
        preserved_values = preserved_state.pop(0) if preserved_state else {}
        line_item = LineItem.objects.create(
            invoice=invoice,
            item_type=item_type,
            job=resolve_job(
                vendor,
                _line_item_job_id(line_item_payload),
                _line_item_job_name(line_item_payload),
            ),
            item_id=str(line_item_payload.get('id') or ''),
            name=str(line_item_payload.get('name') or ''),
            description=str(line_item_payload.get('description') or ''),
            qty=_line_item_qty(line_item_payload.get('qty')),
            unit=str(line_item_payload.get('unit') or ''),
            unit_price=_decimal_12_4(line_item_payload.get('unit_price')) or Decimal('0'),
            total_price=_decimal_12_4(line_item_payload.get('total_price')) or Decimal('0'),
            width=_decimal_12_4(line_item_payload.get('width')),
            length=_decimal_12_4(line_item_payload.get('length')),
            height=_decimal_12_4(line_item_payload.get('height')),
            received=bool(preserved_values.get('received', False)),
            notes=str(preserved_values.get('notes') or ''),
            raw_data=line_item_payload,
        )
        _update_inventory_from_line_item(invoice, line_item, line_item_payload)


@transaction.atomic
def upsert_invoice_from_payload(message_id, email_payload, invoice_payload, vendor):
    """Create or update Invoice and related line items from one parsed invoice dict."""
    email_payload = email_payload or {}
    source_email_date = _parse_datetime(email_payload.get('date'))
    contact = resolve_contact(vendor, email_payload)
    if not contact and vendor:
        contact = vendor.contacts.filter(is_primary=True).first()
    existing_invoice = (
        Invoice.objects.filter(source_email_id=message_id)
        .prefetch_related('line_items__job')
        .first()
    )
    existing_line_item_state = _line_item_state_map(existing_invoice) if existing_invoice else {}
    existing_received_at = existing_invoice.received_at if existing_invoice else None

    defaults = {
        'vendor': vendor,
        'contact': contact,
        'source_email_subject': email_payload.get('subject') or '',
        'source_email_from': email_payload.get('from') or '',
        'source_email_date': source_email_date,
        'received_at': None,
        'invoice_number': str(invoice_payload.get('invoice_number') or ''),
        'invoice_date': _parse_date(invoice_payload.get('date_ordered')),
        'ship_date': _parse_date(invoice_payload.get('ship_date')),
        'due_date': _parse_date(invoice_payload.get('invoice_due_date')),
        'customer_po': _invoice_customer_po(invoice_payload),
        'invoice_total': _decimal_12_2(invoice_payload.get('invoice_total')),
        'status': 'processed',
        'processed_at': timezone.now(),
        'raw_data': invoice_payload,
    }
    invoice, created = Invoice.objects.update_or_create(
        source_email_id=message_id,
        defaults=defaults,
    )
    if not created and existing_received_at != invoice.received_at:
        invoice.received_at = existing_received_at
        invoice.save(update_fields=['received_at', 'updated_at'])
    if not created:
        _revert_inventory_from_existing_invoice(existing_invoice)
        invoice.line_items.all().delete()
    _create_line_items_for_invoice(invoice, vendor, invoice_payload, existing_state=existing_line_item_state)
    sync_invoice_receipt_status(invoice)
    return invoice


@transaction.atomic
def persist_parsed_invoices(vendor, email_payload, parsed_output, message_id_base):
    """
    Persist normalized parser output to Invoice, LineItem, Job, Contact, and InventoryItem.

    ``parsed_output`` is ``{vendor_name, invoices: [...]}`` from ``normalize_parser_output``.
    """
    parsed_output = parsed_output or {}
    email_payload = email_payload or {}
    vendor = resolve_vendor_for_persist(
        vendor=vendor,
        parsed_output=parsed_output,
        email_payload=email_payload,
    )
    invoices_data = parsed_output.get('invoices') or []
    if not invoices_data:
        return []

    saved = []
    for index, invoice_payload in enumerate(invoices_data, start=1):
        source_id = f'{message_id_base}:{index}'
        saved.append(
            upsert_invoice_from_payload(
                source_id,
                email_payload,
                invoice_payload,
                vendor,
            )
        )
    return saved


def save_parsed_invoice(message_id, email_payload, invoice_payload, vendor):
    """Backward-compatible alias for a single parsed invoice."""
    return upsert_invoice_from_payload(message_id, email_payload, invoice_payload, vendor)


def _line_item_to_parser_dict(line_item):
    job = line_item.job if line_item.job_id else None
    return {
        'id': line_item.item_id,
        'name': line_item.name,
        'description': line_item.description,
        'qty': str(line_item.qty),
        'unit': line_item.unit,
        'unit_price': float(line_item.unit_price),
        'total_price': float(line_item.total_price),
        'width': float(line_item.width) if line_item.width is not None else None,
        'length': float(line_item.length) if line_item.length is not None else None,
        'height': float(line_item.height) if line_item.height is not None else None,
        'job_id': job.job_id if job else '',
        'job': job.name if job else '',
        'received': line_item.received,
        'notes': line_item.notes,
    }


def invoice_to_parser_dict(invoice):
    """Map a saved Invoice to the parser-shaped dict used by the processing dialog."""
    if hasattr(invoice, 'line_items'):
        line_items = [
            _line_item_to_parser_dict(line_item)
            for line_item in invoice.line_items.select_related('job').all()
        ]
    else:
        line_items = []
        for line_item in (invoice.get('line_items') or []):
            line_items.append({
                'id': line_item.get('item_id') or line_item.get('id') or '',
                'name': line_item.get('name') or '',
                'description': line_item.get('description') or '',
                'qty': str(line_item.get('qty') or '1'),
                'unit': line_item.get('unit') or '',
                'unit_price': float(line_item.get('unit_price') or 0),
                'total_price': float(line_item.get('total_price') or 0),
                'width': line_item.get('width'),
                'length': line_item.get('length'),
                'height': line_item.get('height'),
                'job_id': line_item.get('job_number') or line_item.get('job_id') or '',
                'job': line_item.get('job_name') or line_item.get('job') or '',
                'received': bool(line_item.get('received', False)),
                'notes': line_item.get('notes') or '',
            })

    def _date_str(value):
        if value is None or value == '':
            return None
        if hasattr(value, 'isoformat'):
            return value.isoformat()
        return str(value)

    vendor_name = ''
    if hasattr(invoice, 'vendor') and invoice.vendor_id:
        vendor_name = invoice.vendor.name
    elif isinstance(invoice, dict):
        vendor_name = invoice.get('vendor_name') or ''

    if isinstance(invoice, dict):
        return {
            'invoice_number': invoice.get('invoice_number') or '',
            'vendor_name': vendor_name,
            'ship_date': _date_str(invoice.get('ship_date')),
            'date_ordered': _date_str(invoice.get('invoice_date') or invoice.get('date_ordered')),
            'invoice_due_date': _date_str(invoice.get('due_date') or invoice.get('invoice_due_date')),
            'cust_po': invoice.get('customer_po') or invoice.get('cust_po') or '',
            'invoice_total': (
                str(invoice['invoice_total'])
                if invoice.get('invoice_total') is not None
                else None
            ),
            'line_items': line_items,
        }

    return {
        'invoice_number': invoice.invoice_number,
        'vendor_name': vendor_name,
        'ship_date': _date_str(invoice.ship_date),
        'date_ordered': _date_str(invoice.invoice_date),
        'invoice_due_date': _date_str(invoice.due_date),
        'cust_po': invoice.customer_po,
        'invoice_total': (
            str(invoice.invoice_total) if invoice.invoice_total is not None else None
        ),
        'line_items': line_items,
    }


def parsed_envelope_for_process_result(result, email_id=None, vendor=None):
    """
    Build ``{vendor_name, invoices}`` for the processing dialog from parse results
    or saved Invoice rows.
    """
    parsed = result.get('parsed') if isinstance(result.get('parsed'), dict) else {}
    invoices = parsed.get('invoices') if parsed else None
    if invoices:
        vendor_name = parsed.get('vendor_name') or (vendor.name if vendor else '')
        return {'vendor_name': vendor_name, 'invoices': list(invoices)}

    # Legacy ProcessedEmail.data may store a single invoice dict.
    if parsed and (parsed.get('line_items') is not None or parsed.get('invoice_number')):
        vendor_name = parsed.get('vendor_name') or (vendor.name if vendor else '')
        return {'vendor_name': vendor_name, 'invoices': [parsed]}

    saved = result.get('invoices') or []
    if saved:
        if saved and hasattr(saved[0], 'pk'):
            invoice_ids = [invoice.pk for invoice in saved]
            saved = list(
                Invoice.objects.filter(pk__in=invoice_ids)
                .select_related('vendor')
                .prefetch_related('line_items__job')
                .order_by('source_email_id')
            )
        vendor_name = vendor.name if vendor else ''
        invoice_dicts = [invoice_to_parser_dict(invoice) for invoice in saved]
        if not vendor_name and invoice_dicts:
            vendor_name = invoice_dicts[0].get('vendor_name') or ''
        return {'vendor_name': vendor_name, 'invoices': invoice_dicts}

    if email_id:
        db_invoices = list(
            Invoice.objects.filter(
                Q(source_email_id=email_id)
                | Q(source_email_id__startswith=f'{email_id}:')
            )
            .select_related('vendor')
            .prefetch_related('line_items__job')
            .order_by('source_email_id')
        )
        if db_invoices:
            if not vendor and db_invoices[0].vendor_id:
                vendor = db_invoices[0].vendor
            vendor_name = vendor.name if vendor else (db_invoices[0].vendor.name if db_invoices[0].vendor_id else '')
            return {
                'vendor_name': vendor_name,
                'invoices': [invoice_to_parser_dict(invoice) for invoice in db_invoices],
            }

    return {'vendor_name': vendor.name if vendor else '', 'invoices': []}


def _list_message_ids(service, query):
    page_token = None
    while True:
        response = service.users().messages().list(
            userId='me',
            q=query,
            maxResults=100,
            pageToken=page_token,
        ).execute()
        for message in response.get('messages', []):
            yield message['id']
        page_token = response.get('nextPageToken')
        if not page_token:
            break


def _select_attachment_part(payload_parts):
    for part in payload_parts or []:
        if part.get('filename') and part.get('mimeType', '').lower() in {'application/pdf', 'application/octet-stream'}:
            return part
    return None


def process_gmail_message(service, message_id):
    existing = ProcessedEmail.objects.filter(email_id=message_id).first()
    if existing and existing.status in ('processed', 'incorrect_parsing'):
        return {
            'status': 'skipped',
            'reason': f'already {existing.status}',
            'processed_email': existing,
        }

    email = service.users().messages().get(userId='me', id=message_id).execute()
    headers = email.get('payload', {}).get('headers', [])
    from_header = next((header['value'] for header in headers if header['name'].lower() == 'from'), '')
    subject = next((header['value'] for header in headers if header['name'].lower() == 'subject'), '')
    date_header = next((header['value'] for header in headers if header['name'].lower() == 'date'), '')
    sender_email = _extract_sender_email(from_header)
    vendor = _sync_vendor_for_sender(from_header, sender_email) if sender_email else None
    if vendor_is_ignored(vendor):
        return {'status': 'skipped', 'reason': 'vendor ignored'}
    payload = email.get('payload', {})
    attachment = _select_attachment_part(payload.get('parts', []))
    if not attachment:
        processed_email, _ = ProcessedEmail.objects.update_or_create(
            email_id=message_id,
            defaults={
                'status': 'error',
                'processed': timezone.now(),
                'data': {'error': 'No PDF attachment found', 'subject': subject},
                'vendor': vendor,
            },
        )
        return {'status': 'error', 'reason': 'no pdf attachment', 'processed_email': processed_email}

    attachment_payload = service.users().messages().attachments().get(
        userId='me',
        messageId=message_id,
        id=attachment['body']['attachmentId'],
    ).execute()

    from base64 import urlsafe_b64decode
    media_dir = settings.MEDIA_ROOT
    os.makedirs(media_dir, exist_ok=True)
    safe_filename = re.sub(r'[^a-zA-Z0-9.-]', '_', attachment['filename'])
    stored_filename = f'{message_id}_{safe_filename}'
    file_path = os.path.join(media_dir, stored_filename)
    with open(file_path, 'wb') as handle:
        handle.write(urlsafe_b64decode(attachment_payload['data'].encode('UTF-8')))

    attachment_info = {
        'filename': stored_filename,
        'original_filename': attachment['filename'],
        'mimeType': attachment.get('mimeType', 'application/pdf'),
        'url': media_url_for_stored_filename(stored_filename),
    }
    _update_email_cache_attachment(message_id, attachment_info)

    parser = _selected_parser_for_vendor(vendor)
    if not parser:
        processed_email, _ = ProcessedEmail.objects.update_or_create(
            email_id=message_id,
            defaults={
                'status': 'error',
                'processed': timezone.now(),
                'data': {'error': 'No parser configured for vendor', 'subject': subject},
                'vendor': vendor,
                'invoice': None,
            },
        )
        return {
            'status': 'error',
            'reason': 'no parser configured',
            'processed_email': processed_email,
            'attachment': attachment_info,
        }

    parsed = normalize_parser_output(
        parser(file_path),
        vendor_name=getattr(parser, 'name', None) or (vendor.name if vendor else None),
    )
    email_payload = {'from': from_header, 'subject': subject, 'date': date_header}
    created_invoices = persist_parsed_invoices(
        vendor,
        email_payload,
        parsed,
        message_id,
    )
    stored_filename = _rename_attachment_for_invoices(
        file_path,
        message_id,
        vendor,
        created_invoices,
        attachment['filename'],
    )
    attachment_info = {
        **attachment_info,
        'filename': stored_filename,
        'url': media_url_for_stored_filename(stored_filename),
    }
    _update_email_cache_attachment(message_id, attachment_info)

    processed_email, _ = ProcessedEmail.objects.update_or_create(
        email_id=message_id,
        defaults={
            'status': 'processed',
            'processed': timezone.now(),
            'data': parsed,
            'vendor': vendor,
            'invoice': created_invoices[0] if created_invoices else None,
        },
    )
    return {
        'status': 'processed',
        'processed_email': processed_email,
        'invoices': created_invoices,
        'parsed': parsed,
        'attachment': attachment_info,
    }


def process_pending_gmail_invoices(limit=None):
    settings_obj = _ensure_invoice_automation_settings()
    if not settings_obj.auto_process_enabled:
        return {'status': 'disabled', 'processed': 0}

    service = get_gmail_service()
    cutoff = timezone.now() - timedelta(days=settings_obj.max_email_age_days)
    query = f"{GMAIL_INVOICE_QUERY} after:{cutoff.strftime('%Y/%m/%d')}"

    processed = 0
    results = []
    interrupted = False
    for message_id in _list_message_ids(service, query):
        if limit is not None and processed >= limit:
            break
        try:
            settings_obj.refresh_from_db(fields=['auto_process_enabled', 'last_processed_at'])
        except InvoiceAutomationSettings.DoesNotExist:
            interrupted = True
            break
        if not settings_obj.auto_process_enabled:
            interrupted = True
            break
        if ProcessedEmail.objects.filter(
            email_id=message_id,
            status__in=('processed', 'incorrect_parsing'),
        ).exists():
            continue
        try:
            result = process_gmail_message(service, message_id)
            if result.get('status') == 'processed':
                processed += 1
            results.append(result)
        except Exception as exc:
            logger.exception('Error auto-processing message %s', message_id)
            ProcessedEmail.objects.update_or_create(
                email_id=message_id,
                defaults={
                    'status': 'error',
                    'processed': timezone.now(),
                    'data': {'error': str(exc)},
                },
            )

    if not interrupted:
        settings_obj.last_processed_at = timezone.now()
        settings_obj.save(update_fields=['last_processed_at', 'updated_at'])
    return {'status': 'ok', 'processed': processed, 'results': results}


def get_automation_settings():
    return _ensure_invoice_automation_settings()


def update_automation_settings(**kwargs):
    settings_obj = _ensure_invoice_automation_settings()
    for field in ('auto_process_enabled', 'max_email_age_days', 'poll_interval_seconds'):
        if field in kwargs and kwargs[field] is not None:
            setattr(settings_obj, field, kwargs[field])
    settings_obj.save()
    return settings_obj


def export_invoices_workbook():
    workbook = Workbook()
    invoice_sheet = workbook.active
    invoice_sheet.title = 'Invoices'
    line_sheet = workbook.create_sheet('Line Items')
    inventory_sheet = workbook.create_sheet('Inventory')

    invoice_headers = [
        'Invoice ID', 'Source Email ID', 'Vendor', 'Contact', 'Invoice Number', 'Invoice Date',
        'Received At', 'Ship Date', 'Due Date', 'Customer PO', 'Invoice Total', 'Status',
        'Processed At', 'Source Email From', 'Source Email Subject',
    ]
    line_headers = [
        'Invoice ID', 'Invoice Number', 'Item Type', 'Item ID', 'Name', 'Description',
        'Job ID', 'Job Name',
        'Qty', 'Unit', 'Unit Price', 'Total Price', 'Width', 'Length', 'Height',
        'Received', 'Notes',
    ]
    inventory_headers = [
        'Item ID', 'Vendor', 'Item Type', 'Item Key', 'Name', 'Description', 'Unit', 'Current Qty',
        'Last Unit Price', 'Last Total Price', 'Last Invoiced At',
    ]

    invoice_sheet.append(invoice_headers)
    line_sheet.append(line_headers)
    inventory_sheet.append(inventory_headers)

    invoice_qs = exclude_ignored_vendor_relations(
        Invoice.objects.select_related('vendor').prefetch_related('line_items')
    ).order_by('-received_at', '-processed_at', '-created_at')
    for invoice in invoice_qs:
        invoice_sheet.append([
            invoice.id,
            invoice.source_email_id,
            invoice.vendor.name if invoice.vendor else '',
            invoice.contact.name if invoice.contact else '',
            invoice.invoice_number,
            invoice.invoice_date.isoformat() if invoice.invoice_date else '',
            invoice.received_at.isoformat() if invoice.received_at else '',
            invoice.ship_date.isoformat() if invoice.ship_date else '',
            invoice.due_date.isoformat() if invoice.due_date else '',
            invoice.customer_po,
            float(invoice.invoice_total) if invoice.invoice_total is not None else '',
            invoice.status,
            invoice.processed_at.isoformat() if invoice.processed_at else '',
            invoice.source_email_from,
            invoice.source_email_subject,
        ])

        for line_item in invoice.line_items.all():
            line_sheet.append([
                invoice.id,
                invoice.invoice_number,
                line_item.item_type.get_full_path() if line_item.item_type else '',
                line_item.item_id,
                line_item.name,
                line_item.description,
                line_item.job.job_id if line_item.job_id else '',
                line_item.job.name if line_item.job_id else '',
                float(line_item.qty),
                line_item.unit,
                float(line_item.unit_price),
                float(line_item.total_price),
                float(line_item.width) if line_item.width is not None else '',
                float(line_item.length) if line_item.length is not None else '',
                float(line_item.height) if line_item.height is not None else '',
                'Yes' if line_item.received else 'No',
                line_item.notes,
            ])

    for item in InventoryItem.objects.select_related('vendor').order_by('name', 'item_key'):
        inventory_sheet.append([
            item.id,
            item.vendor.name if item.vendor else '',
            item.item_type.get_full_path() if item.item_type else '',
            item.item_key,
            item.name,
            item.description,
            item.unit,
            float(item.current_qty),
            float(item.last_unit_price) if item.last_unit_price is not None else '',
            float(item.last_total_price) if item.last_total_price is not None else '',
            item.last_invoiced_at.isoformat() if item.last_invoiced_at else '',
        ])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def gmail_message_id_from_source_email_id(source_email_id):
    """Gmail message id from invoice ``source_email_id`` (may include ``:index`` suffix)."""
    if not source_email_id:
        return ''
    return source_email_id.split(':', 1)[0]


def reset_processed_email_after_invoice_deleted(invoice):
    """
    When the last invoice from a Gmail message is removed, reset ProcessedEmail so the
    email can be imported again.
    """
    message_id = gmail_message_id_from_source_email_id(invoice.source_email_id)
    if not message_id:
        return

    still_linked = Invoice.objects.filter(
        Q(source_email_id=message_id) | Q(source_email_id__startswith=f'{message_id}:')
    ).exclude(pk=invoice.pk).exists()
    if still_linked:
        return

    ProcessedEmail.objects.filter(email_id=message_id).update(
        status='pending',
        processed=None,
        invoice=None,
    )


@transaction.atomic
def reset_invoice_data(remove_all=False):
    """
    Clear imported invoice data.

    By default this preserves vendor and item-type setup, while removing imported
    email/cache/contact/job/invoice/inventory state. ``remove_all=True`` also
    removes vendor, item-type, and automation settings rows.
    """
    media_dir = settings.MEDIA_ROOT
    attachment_paths = []
    if os.path.isdir(media_dir):
        for root, _dirs, files in os.walk(media_dir):
            for name in files:
                if name.lower().endswith('.pdf'):
                    attachment_paths.append(os.path.join(root, name))

    counts = {
        'email_message_cache': EmailMessageCache.objects.count(),
        'processed_emails': ProcessedEmail.objects.count(),
        'line_items': LineItem.objects.count(),
        'invoices': Invoice.objects.count(),
        'inventory_items': InventoryItem.objects.count(),
        'contacts': Contact.objects.count(),
        'jobs': Job.objects.count(),
        'vendor_emails': VendorEmail.objects.count(),
    }
    if remove_all:
        counts.update({
            'vendors': Vendor.objects.count(),
            'item_types': ItemType.objects.count(),
            'automation_settings': InvoiceAutomationSettings.objects.count(),
        })

    # Use SQL bulk deletes so rows with out-of-range decimals (e.g. mis-parsed
    # phone numbers stored as qty) do not need to hydrate through the ORM.
    db = LineItem.objects.db
    LineItem.objects.all()._raw_delete(using=db)
    ProcessedEmail.objects.all()._raw_delete(using=db)
    Invoice.objects.all()._raw_delete(using=db)
    InventoryItem.objects.all()._raw_delete(using=db)
    EmailMessageCache.objects.all()._raw_delete(using=db)
    Contact.objects.all()._raw_delete(using=db)
    Job.objects.all()._raw_delete(using=db)
    VendorEmail.objects.all()._raw_delete(using=db)

    if remove_all:
        Vendor.objects.all()._raw_delete(using=db)
        ItemType.objects.all()._raw_delete(using=db)
        InvoiceAutomationSettings.objects.all()._raw_delete(using=db)

    deleted_files = 0
    for file_path in attachment_paths:
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                deleted_files += 1
        except OSError:
            logger.exception('Failed to delete attachment %s', file_path)

    return {
        'deleted_counts': counts,
        'deleted_files': deleted_files,
        'remove_all': bool(remove_all),
    }


def start_autoprocess_worker():
    global _worker_thread
    with _worker_lock:
        if _worker_thread and _worker_thread.is_alive():
            return _worker_thread

        def _run():
            while True:
                poll_interval = 60
                try:
                    settings_obj = _ensure_invoice_automation_settings()
                    poll_interval = max(15, settings_obj.poll_interval_seconds)
                    if settings_obj.auto_process_enabled:
                        if _processing_lock.acquire(blocking=False):
                            try:
                                process_pending_gmail_invoices()
                            finally:
                                _processing_lock.release()
                except Exception:
                    logger.exception('Auto-processing worker failed')
                time.sleep(poll_interval)

        _worker_thread = threading.Thread(target=_run, name='invoiceinator-autoprocess', daemon=True)
        _worker_thread.start()
        return _worker_thread
